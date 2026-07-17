"""
Report Sales Order lines and Service (Customer Service) lines whose unit differs
from the item's default (stock) unit AND has no usable UnitConversion configured
— i.e. the lines that would raise "No unit conversion configured …" when their
quantity is converted to the inventory unit.

Scope is deliberately limited to Sales Order lines and Service lines. It
EXCLUDES deliveries, pickups, and procurement (PO/GRN) — those are downstream/
inbound documents handled elsewhere (e.g. resync_inventory covers posted stock
movements).

Mirrors the check convert_to_base_unit() itself performs (item-specific first,
then global; direct or reverse), so a line is only reported when conversion
would genuinely fail.

Usage:
    python manage.py report_missing_conversions
    python manage.py report_missing_conversions --all
    python manage.py report_missing_conversions --excel
    python manage.py report_missing_conversions --excel "C:/path/report.xlsx"
"""
import datetime

from django.core.management.base import BaseCommand

from catalog.models import convert_to_base_unit

DEFAULT_XLSX = 'missing_conversions_latest.xlsx'

# Scope → which source collectors to run. 'service' is intentionally in BOTH
# the order scope (it's manually entered, like a Sales Order) and the stock
# scope (a posted Customer Service actually deducts stock). The union in 'all'
# runs each collector once (see collect_problematic_lines).
#   order  — what the human INTENDED: Sales Order + Service lines. The
#            authoritative catalog of conversions the business means to use.
#   stock  — what actually HITS inventory: posted Delivery/Pickup/GRN/POS +
#            Service lines. This is the set resync_inventory enforces, so a
#            pair here that is NOT in 'order' means a document diverged from
#            any order (edited/no-SO/POS/bug) — the resync-blocking cases an
#            order-only report can't see.
#   all    — union of both.
ORDER_SOURCES = ('sales_order', 'service')
STOCK_SOURCES = ('delivery', 'pickup', 'grn', 'pos', 'service')


def _missing_conversion(line):
    """True when line.unit != item.default_unit and no conversion is available."""
    item, used, default = line.item, line.unit, line.item.default_unit
    if used.pk == default.pk:
        return False
    try:
        convert_to_base_unit(1, used, default, item)
        return False
    except Exception:
        return True


def _row(line, source, doc, date, qty, price):
    return {
        'name': line.item.name, 'code': line.item.code,
        'default': line.item.default_unit.abbreviation,
        'used': line.unit.abbreviation,
        'source': source, 'doc': doc, 'date': date, 'qty': qty, 'price': price,
    }


def _rows_sales_order():
    from sales.models import SalesOrderLine
    qs = SalesOrderLine.objects.select_related(
        'item__default_unit', 'unit', 'sales_order'
    ).filter(sales_order__is_active=True)  # exclude soft-deleted SOs
    return [_row(l, 'Sales Order Line', l.sales_order.document_number,
                 l.sales_order.order_date, l.qty_ordered, l.unit_price)
            for l in qs if _missing_conversion(l)]


def _rows_service():
    from services.models import ServiceLine
    qs = ServiceLine.objects.select_related('item__default_unit', 'unit', 'service')
    return [_row(l, 'Service Line', l.service.service_number,
                 l.service.service_date, l.qty, l.unit_price)
            for l in qs if _missing_conversion(l)]


def _rows_delivery():
    from sales.models import DeliveryLine
    from core.models import DocumentStatus
    qs = DeliveryLine.objects.select_related(
        'item__default_unit', 'unit', 'delivery'
    ).filter(delivery__status=DocumentStatus.POSTED, delivery__is_active=True)
    return [_row(l, 'Delivery Line', l.delivery.document_number,
                 l.delivery.delivery_date, l.qty, None)
            for l in qs if _missing_conversion(l)]


def _rows_pickup():
    from sales.models import SalesPickupLine
    from core.models import DocumentStatus
    qs = SalesPickupLine.objects.select_related(
        'item__default_unit', 'unit', 'pickup'
    ).filter(pickup__status=DocumentStatus.POSTED, pickup__is_active=True)
    return [_row(l, 'Pickup Line', l.pickup.document_number,
                 l.pickup.pickup_date, l.qty, None)
            for l in qs if _missing_conversion(l)]


def _rows_grn():
    from procurement.models import GoodsReceiptLine
    from core.models import DocumentStatus
    qs = GoodsReceiptLine.objects.select_related(
        'item__default_unit', 'unit', 'goods_receipt'
    ).filter(goods_receipt__status=DocumentStatus.POSTED, goods_receipt__is_active=True)
    return [_row(l, 'GRN Line', l.goods_receipt.document_number,
                 l.goods_receipt.receipt_date, l.qty, None)
            for l in qs if _missing_conversion(l)]


def _rows_pos():
    from pos.models import POSSaleLine, SaleStatus
    qs = POSSaleLine.objects.select_related(
        'item__default_unit', 'unit', 'sale'
    ).filter(sale__status=SaleStatus.POSTED)
    return [_row(l, 'POS Line', l.sale.sale_no,
                 l.sale.created_at.date() if l.sale.created_at else None,
                 l.qty, l.unit_price)
            for l in qs if _missing_conversion(l)]


_SOURCE_FUNCS = {
    'sales_order': _rows_sales_order,
    'service': _rows_service,
    'delivery': _rows_delivery,
    'pickup': _rows_pickup,
    'grn': _rows_grn,
    'pos': _rows_pos,
}


def collect_problematic_lines(scope='order'):
    """Return problematic-line dicts for the given scope, newest document first.

    scope: 'order' (Sales Order + Service), 'stock' (posted Delivery/Pickup/
    GRN/POS + Service — what resync_inventory enforces), or 'all' (union).
    Each source collector runs at most once even when a scope lists it twice
    (e.g. 'service' is in both order and stock).
    """
    names = []
    if scope in ('order', 'all'):
        names += ORDER_SOURCES
    if scope in ('stock', 'all'):
        names += STOCK_SOURCES

    rows, seen = [], set()
    for name in names:
        if name in seen:
            continue
        seen.add(name)
        rows += _SOURCE_FUNCS[name]()

    rows.sort(key=lambda r: (r['date'] or datetime.date.min), reverse=True)
    return rows


def latest_per_pair(rows):
    """Collapse to the latest line per (item, used->default), with a count.

    Returns (latest_rows_newest_first, counts) where counts is keyed by
    (code, used, default).
    """
    latest, counts = {}, {}
    for r in rows:  # rows already newest-first
        key = (r['code'], r['used'], r['default'])
        counts[key] = counts.get(key, 0) + 1
        latest.setdefault(key, r)
    latest_rows = sorted(
        latest.values(),
        key=lambda r: (r['date'] or datetime.date.min), reverse=True)
    return latest_rows, counts


class Command(BaseCommand):
    help = ("Report lines whose unit has no conversion to the item's default "
            'unit. --scope order (default) = Sales Order + Service (intent); '
            'stock = posted Delivery/Pickup/GRN/POS + Service (what resync '
            'enforces); all = union.')

    def add_arguments(self, parser):
        parser.add_argument(
            '--scope', choices=('order', 'stock', 'all'), default='order',
            help="'order' (Sales Order + Service, default), 'stock' (posted "
                 "Delivery/Pickup/GRN/POS + Service), or 'all' (union).",
        )
        parser.add_argument(
            '--all', action='store_true',
            help='List every problematic line, not just the latest per item/unit.',
        )
        parser.add_argument(
            '--excel', nargs='?', const=DEFAULT_XLSX, default=None, metavar='PATH',
            help=f'Also write an .xlsx report (default filename: {DEFAULT_XLSX}).',
        )

    def handle(self, *args, **options):
        scope = options['scope']
        rows = collect_problematic_lines(scope=scope)
        latest_rows, counts = latest_per_pair(rows)

        self.stdout.write(self.style.MIGRATE_HEADING(
            f'MISSING UNIT CONVERSIONS  [scope={scope}]'))
        if not rows:
            self.stdout.write(self.style.SUCCESS(
                '\n✓ No problematic lines — every non-default unit has a conversion.'))
            return

        self.stdout.write(
            f'\nFound {len(rows)} problematic line(s) across '
            f'{len(latest_rows)} distinct item/unit conversion(s).\n')

        # Latest per distinct missing conversion (the actionable list)
        self.stdout.write(self.style.WARNING(
            'Latest occurrence per missing conversion:'))
        for r in latest_rows:
            key = (r['code'], r['used'], r['default'])
            self.stdout.write(
                f"  {r['code']:<28} {r['used']:>5} -> {r['default']:<5} "
                f"| add: 1 {r['used']} = ? {r['default']:<5} "
                f"| latest {r['source']} {r['doc']} ({r['date']}) "
                f"| {counts[key]}x"
            )

        if options['all']:
            self.stdout.write('\n' + self.style.WARNING('All problematic lines:'))
            for r in rows:
                price = '' if r['price'] is None else f"{r['price']}"
                self.stdout.write(
                    f"  {r['date']} {r['source']:<8} {r['doc']:<14} "
                    f"{r['code']:<28} qty {r['qty']} {r['used']} "
                    f"(default {r['default']}) {price}"
                )

        if options['excel'] is not None:
            path = self._write_excel(options['excel'], rows, latest_rows, counts)
            self.stdout.write('\n' + self.style.SUCCESS(f'Excel written: {path}'))

    # ── Excel export ────────────────────────────────────────────────────
    def _write_excel(self, path, rows, latest_rows, counts):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        head_fill = PatternFill('solid', fgColor='C0392B')
        head_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def iso(d):
            return d.isoformat() if d else ''

        def style_header(ws, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill, cell.font, cell.alignment, cell.border = (
                    head_fill, head_font, center, border)
            ws.freeze_panes = 'A2'

        def autosize(ws, headers):
            for i in range(1, len(headers) + 1):
                width = len(str(headers[i - 1])) + 2
                for row in ws.iter_rows(min_col=i, max_col=i, min_row=2):
                    v = row[0].value
                    if v is not None:
                        width = max(width, len(str(v)) + 2)
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 48)

        def fnum(v):
            return float(v) if v is not None else None

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = 'Latest Missing Conversions'
        h1 = ['Item Name', 'Item Code', 'Default Unit', 'Unit Used',
              'Needed Conversion', 'Source', 'Latest Document No.', 'Latest Date',
              'Qty', 'Unit Price', 'Occurrences']
        ws1.append(h1)
        for r in latest_rows:
            key = (r['code'], r['used'], r['default'])
            ws1.append([
                r['name'], r['code'], r['default'], r['used'],
                f"1 {r['used']} = ? {r['default']}", r['source'], r['doc'],
                iso(r['date']), fnum(r['qty']), fnum(r['price']), counts[key],
            ])
        style_header(ws1, len(h1))
        autosize(ws1, h1)

        ws2 = wb.create_sheet('All Problematic Lines')
        h2 = ['Item Name', 'Item Code', 'Default Unit', 'Unit Used',
              'Needed Conversion', 'Source', 'Document No.', 'Date', 'Qty', 'Unit Price']
        ws2.append(h2)
        for r in rows:
            ws2.append([
                r['name'], r['code'], r['default'], r['used'],
                f"1 {r['used']} = ? {r['default']}", r['source'], r['doc'],
                iso(r['date']), fnum(r['qty']), fnum(r['price']),
            ])
        style_header(ws2, len(h2))
        autosize(ws2, h2)

        wb.save(path)
        return path
