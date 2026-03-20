from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import ProtectedError
from django.core.paginator import Paginator
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from catalog.models import Category, Unit, UnitConversion, Item
from core.utils import (
    build_relation_summary,
    handle_delete_error,
    handle_delete_success,
    handle_protected_error,
    format_in_use_message,
)
from catalog.serializers import (
    CategorySerializer, UnitSerializer, UnitConversionSerializer,
    ItemSerializer, ItemListSerializer,
)
from catalog.forms import CategoryForm, UnitForm, UnitConversionForm, ItemForm, ItemUnitConversionFormSet
from inventory.models import StockBalance
from django.db.models import Sum, F
from inventory.serializers import StockBalanceSerializer


# ── API Views ──────────────────────────────────────────────────────────────

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    search_fields = ['name', 'code']
    filterset_fields = ['parent', 'is_active']


class UnitViewSet(viewsets.ModelViewSet):
    queryset = Unit.objects.all()
    serializer_class = UnitSerializer
    search_fields = ['name', 'abbreviation']
    filterset_fields = ['category', 'is_active']


class UnitConversionViewSet(viewsets.ModelViewSet):
    queryset = UnitConversion.objects.select_related('from_unit', 'to_unit').all()
    serializer_class = UnitConversionSerializer


class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.select_related('category', 'default_unit').all()
    search_fields = ['code', 'name', 'barcode']
    filterset_fields = ['item_type', 'category', 'is_active']

    def get_serializer_class(self):
        if self.action == 'list':
            return ItemListSerializer
        return ItemSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Optional available stock map for a specific register/warehouse
        register_id = request.query_params.get('register')
        available_map = {}
        if register_id:
            from pos.models import POSRegister
            register = POSRegister.objects.filter(pk=register_id).select_related('warehouse').first()
            if register:
                balances = StockBalance.objects.filter(
                    location__warehouse=register.warehouse
                ).values('item').annotate(
                    available=Sum(F('qty_on_hand') - F('qty_reserved'))
                )
                available_map = {row['item']: row['available'] for row in balances}

        page = self.paginate_queryset(queryset)
        serializer_kwargs = {'context': self.get_serializer_context()}
        serializer_kwargs['context']['available_map'] = available_map

        if page is not None:
            serializer = self.get_serializer(page, many=True, **serializer_kwargs)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True, **serializer_kwargs)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def stock(self, request, pk=None):
        item = self.get_object()
        balances = StockBalance.objects.filter(
            item=item, qty_on_hand__gt=0
        ).select_related('location', 'location__warehouse')
        serializer = StockBalanceSerializer(balances, many=True)
        return Response(serializer.data)


# ── Template Views ─────────────────────────────────────────────────────────

@login_required
def item_list_view(request):
    items_qs = Item.objects.select_related('category', 'default_unit', 'selling_unit').all()
    item_type = request.GET.get('type', '')
    category_id = request.GET.get('category', '')
    search = request.GET.get('q', '')

    if item_type:
        items_qs = items_qs.filter(item_type=item_type)
    if category_id:
        items_qs = items_qs.filter(category_id=category_id)
    if search:
        items_qs = (
            items_qs.filter(name__icontains=search) |
            items_qs.filter(code__icontains=search)
        )

    items_qs = items_qs.order_by('category__name', 'name')

    try:
        cols = max(1, min(6, int(request.GET.get('cols', 4))))
    except (ValueError, TypeError):
        cols = 4
    try:
        rows = max(1, min(10, int(request.GET.get('rows', 3))))
    except (ValueError, TypeError):
        rows = 3

    per_page = cols * rows
    paginator = Paginator(items_qs, per_page)
    try:
        page_num = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page_num = 1
    page_obj = paginator.get_page(page_num)

    categories = Category.objects.filter(is_active=True).order_by('name')

    params = request.GET.copy()
    params.pop('page', None)
    query_base = params.urlencode()

    return render(request, 'catalog/item_list.html', {
        'page_obj': page_obj,
        'items': page_obj.object_list,
        'categories': categories,
        'current_type': item_type,
        'current_category': category_id,
        'search': search,
        'cols': cols,
        'rows': rows,
        'per_page': per_page,
        'query_base': query_base,
        'col_range': range(1, 7),
        'row_range': range(1, 11),
    })


@login_required
def catalog_export_excel_view(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    items_qs = Item.objects.select_related('category', 'default_unit', 'selling_unit').all()
    item_type = request.GET.get('type', '')
    category_id = request.GET.get('category', '')
    search = request.GET.get('q', '')

    if item_type:
        items_qs = items_qs.filter(item_type=item_type)
    if category_id:
        items_qs = items_qs.filter(category_id=category_id)
    if search:
        items_qs = (
            items_qs.filter(name__icontains=search) |
            items_qs.filter(code__icontains=search)
        )
    items_qs = items_qs.order_by('category__name', 'name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Item Catalog'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    headers = [
        'Code', 'Name', 'Type', 'Category', 'Default Unit', 'Selling Unit',
        'Cost Price', 'Selling Price', 'Barcode',
        'Min Stock', 'Max Stock', 'Reorder Point', 'Description', 'Status',
    ]
    ws.append(headers)
    for idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for item in items_qs:
        ws.append([
            item.code,
            item.name,
            item.get_item_type_display(),
            item.category.name if item.category_id else '',
            item.default_unit.abbreviation if item.default_unit_id else '',
            item.selling_unit.abbreviation if item.selling_unit_id else '',
            float(item.cost_price or 0),
            float(item.selling_price or 0),
            item.barcode or '',
            float(item.minimum_stock or 0),
            float(item.maximum_stock or 0),
            float(item.reorder_point or 0),
            item.description or '',
            'Active' if item.is_active else 'Inactive',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="catalog_items.xlsx"'
    wb.save(response)
    return response


@login_required
def catalog_print_view(request):
    items_qs = Item.objects.select_related('category', 'default_unit', 'selling_unit').all()
    item_type = request.GET.get('type', '')
    category_id = request.GET.get('category', '')
    search = request.GET.get('q', '')

    if item_type:
        items_qs = items_qs.filter(item_type=item_type)
    if category_id:
        items_qs = items_qs.filter(category_id=category_id)
    if search:
        items_qs = (
            items_qs.filter(name__icontains=search) |
            items_qs.filter(code__icontains=search)
        )
    items_qs = items_qs.order_by('category__name', 'name')

    try:
        cols = max(1, min(6, int(request.GET.get('cols', 4))))
    except (ValueError, TypeError):
        cols = 4
    try:
        rows = max(1, min(10, int(request.GET.get('rows', 3))))
    except (ValueError, TypeError):
        rows = 3

    per_page = cols * rows
    items_list = list(items_qs)
    pages = [items_list[i:i + per_page] for i in range(0, max(len(items_list), 1), per_page)]

    return render(request, 'catalog/item_list_print.html', {
        'pages': pages,
        'cols': cols,
        'rows': rows,
        'per_page': per_page,
        'total_count': len(items_list),
    })


@login_required
def item_detail_view(request, pk):
    from decimal import Decimal
    from django.db.models import Sum
    from django.db.models.functions import Coalesce
    item = get_object_or_404(Item, pk=pk)
    item_conversions = item.unit_conversions.select_related('from_unit', 'to_unit').order_by('from_unit__name', 'to_unit__name')
    balances = StockBalance.objects.filter(
        item=item
    ).select_related('location', 'location__warehouse')

    # Aggregate totals per warehouse
    warehouse_summary = balances.values(
        'location__warehouse__code', 'location__warehouse__name'
    ).annotate(
        total_on_hand=Coalesce(Sum('qty_on_hand'), Decimal('0')),
        total_reserved=Coalesce(Sum('qty_reserved'), Decimal('0')),
    ).order_by('location__warehouse__code')

    # Grand totals
    totals = balances.aggregate(
        total_on_hand=Coalesce(Sum('qty_on_hand'), Decimal('0')),
        total_reserved=Coalesce(Sum('qty_reserved'), Decimal('0')),
    )
    totals['total_available'] = totals['total_on_hand'] - totals['total_reserved']
    totals['total_value'] = totals['total_on_hand'] * (item.cost_price or Decimal('0'))

    from inventory.models import StockMove
    recent_moves = StockMove.objects.filter(
        item=item, status='POSTED'
    ).select_related('unit', 'from_location', 'to_location', 'created_by')[:20]
    return render(request, 'catalog/item_detail.html', {
        'item': item,
        'item_conversions': item_conversions,
        'balances': balances,
        'warehouse_summary': warehouse_summary,
        'totals': totals,
        'recent_moves': recent_moves,
    })


@login_required
def item_create_view(request):
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES)
        item = form.instance
        formset = ItemUnitConversionFormSet(request.POST, instance=item, prefix='conversions')
        if form.is_valid() and formset.is_valid():
            item = form.save()
            formset.instance = item
            formset.save()
            messages.success(request, 'Item created successfully.')
            return redirect('item_list')
    else:
        form = ItemForm()
        formset = ItemUnitConversionFormSet(instance=Item(), prefix='conversions')
    return render(request, 'catalog/item_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Create Item',
    })


@login_required
def item_edit_view(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES, instance=item)
        formset = ItemUnitConversionFormSet(request.POST, instance=item, prefix='conversions')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Item updated successfully.')
            return redirect('item_detail', pk=item.pk)
    else:
        form = ItemForm(instance=item)
        formset = ItemUnitConversionFormSet(instance=item, prefix='conversions')
    return render(request, 'catalog/item_form.html', {
        'form': form,
        'formset': formset,
        'title': f'Edit Item: {item.code}',
        'item': item,
    })


@login_required
def item_delete_view(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        item.soft_delete()
        messages.success(request, f'Item {item.code} deleted.')
        return redirect('item_list')
    return render(request, 'catalog/item_delete.html', {'object': item})


# ── Category CRUD ──────────────────────────────────────────────────────────

@login_required
def category_list_view(request):
    categories = Category.objects.all()
    return render(request, 'catalog/category_list.html', {'categories': categories})


@login_required
def category_create_view(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category created successfully.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'catalog/category_form.html', {'form': form, 'title': 'Create Category'})


@login_required
def category_edit_view(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'catalog/category_form.html', {'form': form, 'title': f'Edit Category: {category.name}'})


@login_required
def category_delete_view(request, pk):
    category = get_object_or_404(Category, pk=pk)
    relation_summary = build_relation_summary(category)
    if request.method == 'POST':
        try:
            if category.children.exists() or category.items.exists():
                base_msg = 'Cannot delete category because it is in use or referenced by other records.'
                response = handle_delete_error(request, base_msg, relation_summary)
                if response:
                    return response
            else:
                name = category.name
                category.delete()
                msg = f'Category {name} deleted.'
                response = handle_delete_success(request, msg)
                if response:
                    return response
                return redirect('category_list')
        except ProtectedError as e:
            base_msg = 'Cannot delete category because it is in use or referenced by other records.'
            response = handle_protected_error(request, e, base_msg, relation_summary)
            if response:
                return response
    return render(request, 'catalog/category_delete.html', {'object': category, 'relation_summary': relation_summary})

# ── Unit CRUD ──────────────────────────────────────────────────────────────

@login_required
def unit_list_view(request):
    units = Unit.objects.all()
    return render(request, 'catalog/unit_list.html', {'units': units})

@login_required
def unit_create_view(request):
    if request.method == 'POST':
        form = UnitForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit created successfully.')
            return redirect('unit_list')
    else:
        form = UnitForm()
    return render(request, 'catalog/unit_form.html', {'form': form, 'title': 'Create Unit'})


@login_required
def unit_edit_view(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        form = UnitForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit updated successfully.')
            return redirect('unit_list')
    else:
        form = UnitForm(instance=unit)
    return render(request, 'catalog/unit_form.html', {'form': form, 'title': f'Edit Unit: {unit.name}'})


@login_required
def unit_delete_view(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        unit.soft_delete()
        messages.success(request, f'Unit {unit.name} deleted.')
        return redirect('unit_list')
    return render(request, 'catalog/unit_delete.html', {'object': unit})


# ── Unit Conversion CRUD ─────────────────────────────────────────────

@login_required
def unit_conversion_list_view(request):
    conversions = UnitConversion.objects.select_related(
        'from_unit', 'to_unit', 'item'
    ).all().order_by('from_unit__name', 'to_unit__name')
    return render(request, 'catalog/unit_conversion_list.html', {'conversions': conversions})


@login_required
def unit_conversion_create_view(request):
    if request.method == 'POST':
        form = UnitConversionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit conversion created successfully.')
            return redirect('unit_conversion_list')
    else:
        form = UnitConversionForm()
    return render(request, 'catalog/unit_conversion_form.html', {
        'form': form, 'title': 'Create Unit Conversion'
    })


@login_required
def unit_conversion_edit_view(request, pk):
    conv = get_object_or_404(UnitConversion, pk=pk)
    if request.method == 'POST':
        form = UnitConversionForm(request.POST, instance=conv)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit conversion updated successfully.')
            return redirect('unit_conversion_list')
    else:
        form = UnitConversionForm(instance=conv)
    return render(request, 'catalog/unit_conversion_form.html', {
        'form': form,
        'title': f'Edit: 1 {conv.from_unit.abbreviation} = {conv.factor} {conv.to_unit.abbreviation}'
    })


@login_required
def unit_conversion_delete_view(request, pk):
    conv = get_object_or_404(UnitConversion, pk=pk)
    if request.method == 'POST':
        conv.delete()
        messages.success(request, 'Unit conversion deleted.')
        return redirect('unit_conversion_list')
    return render(request, 'catalog/unit_conversion_delete.html', {'object': conv})
