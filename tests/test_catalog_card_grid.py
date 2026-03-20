"""
Focused regression tests for the catalog card-grid UI:
  - item_list_view  (GET /catalog/items/)
  - catalog_print_view  (GET /catalog/items/print/)
  - catalog_export_excel_view  (GET /catalog/items/export-excel/)
"""
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

User = get_user_model()


class CatalogCardGridViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from catalog.models import Category, Item, ItemType, Unit, UnitCategory

        cls.user = User.objects.create_superuser('catgrid_u', 'cg@test.com', 'pass')
        cls.cat = Category.objects.create(name='Widgets', code='WDGT')
        cls.unit = Unit.objects.create(name='GridPiece', abbreviation='gpc', category=UnitCategory.QUANTITY)

        cls.items = []
        for i in range(15):
            cls.items.append(Item.objects.create(
                code=f'GRID-{i:03d}',
                name=f'Grid Item {i:03d}',
                item_type=ItemType.FINISHED,
                category=cls.cat,
                default_unit=cls.unit,
                cost_price=Decimal('5'),
                selling_price=Decimal('10'),
            ))

    def setUp(self):
        self.client.force_login(self.user)

    # ── item_list_view ────────────────────────────────────────────────────

    def test_list_view_renders_card_grid(self):
        r = self.client.get(reverse('item_list'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'catalog-grid')
        self.assertContains(r, 'catalog-toolbar')

    def test_default_cols_rows(self):
        r = self.client.get(reverse('item_list'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, '--cat-cols: 4')

    def test_custom_cols_rows(self):
        r = self.client.get(reverse('item_list') + '?cols=3&rows=2')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, '--cat-cols: 3')

    def test_pagination_12_per_page_default(self):
        """Default 4 cols × 3 rows = 12 items per page; 15 items → 2 pages."""
        r = self.client.get(reverse('item_list'))
        self.assertEqual(r.status_code, 200)
        page_obj = r.context['page_obj']
        self.assertEqual(page_obj.paginator.num_pages, 2)
        self.assertEqual(len(r.context['items']), 12)

    def test_pagination_page2(self):
        r = self.client.get(reverse('item_list') + '?page=2')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(r.context['items']), 3)

    def test_search_filter(self):
        r = self.client.get(reverse('item_list') + '?q=GRID-005')
        self.assertEqual(r.status_code, 200)
        items = r.context['items']
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].code, 'GRID-005')

    def test_category_filter(self):
        r = self.client.get(reverse('item_list') + f'?category={self.cat.pk}')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.context['page_obj'].paginator.count, 15)

    def test_alphabetical_sort_by_category_then_name(self):
        r = self.client.get(reverse('item_list') + '?cols=6&rows=10')
        items = list(r.context['items'])
        names = [i.name for i in items]
        self.assertEqual(names, sorted(names))

    def test_export_buttons_in_html(self):
        r = self.client.get(reverse('item_list'))
        self.assertContains(r, 'export-pdf-page')
        self.assertContains(r, 'export-pdf-all')
        self.assertContains(r, 'export-excel')

    def test_search_clears_page_to_1(self):
        r = self.client.get(reverse('item_list') + '?q=Grid+Item&page=1')
        self.assertEqual(r.status_code, 200)

    # ── catalog_print_view ────────────────────────────────────────────────

    def test_print_view_returns_200(self):
        r = self.client.get(reverse('catalog_print') + '?cols=4&rows=3')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'print-page')
        self.assertContains(r, 'print-card')

    def test_print_view_splits_pages_correctly(self):
        r = self.client.get(reverse('catalog_print') + '?cols=3&rows=2')
        self.assertEqual(r.status_code, 200)
        pages = r.context['pages']
        self.assertEqual(r.context['cols'], 3)
        self.assertEqual(r.context['rows'], 2)
        self.assertEqual(r.context['per_page'], 6)
        self.assertEqual(r.context['total_count'], 15)
        # 15 items, 6 per page → 3 pages
        self.assertEqual(len(pages), 3)
        self.assertEqual(len(pages[0]), 6)
        self.assertEqual(len(pages[2]), 3)

    def test_print_view_contains_item_names(self):
        r = self.client.get(reverse('catalog_print'))
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn('GRID-000', body)

    # ── catalog_export_excel_view ─────────────────────────────────────────

    def test_excel_export_status_and_content_type(self):
        r = self.client.get(reverse('catalog_export_excel'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_excel_export_content_disposition(self):
        r = self.client.get(reverse('catalog_export_excel'))
        self.assertIn('.xlsx', r.get('Content-Disposition', ''))

    def test_excel_export_contains_all_items(self):
        import openpyxl, io
        r = self.client.get(reverse('catalog_export_excel'))
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Row 1 = header, rows 2+ = data; expect 15 data rows
        self.assertEqual(ws.max_row - 1, 15)

    def test_excel_export_headers(self):
        import openpyxl, io
        r = self.client.get(reverse('catalog_export_excel'))
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn('Code', headers)
        self.assertIn('Name', headers)
        self.assertIn('Selling Price', headers)
        self.assertNotIn('Image', headers)

    def test_excel_export_respects_search_filter(self):
        import openpyxl, io
        r = self.client.get(reverse('catalog_export_excel') + '?q=GRID-005')
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        self.assertEqual(ws.max_row - 1, 1)
        self.assertEqual(ws.cell(row=2, column=1).value, 'GRID-005')
